import requests
import json 
import pandas as pd
import argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Take command-line arguements 
parser = argparse.ArgumentParser()
parser.add_argument('-k', '--keys', nargs='+')      # option that takes a value
parser.add_argument('-c', '--colored', action='store_true')
args = parser.parse_args()

keys = args.keys
colored = args.colored

# Read csv file
pd_csv = pd.read_csv('vehicles.csv', sep=';', encoding='utf-8', keep_default_na=False, na_values=['NaN'])
#  convert the csv data to json data
pd_json = pd_csv.to_json(orient='records', force_ascii=False)


r = requests.post(url="http://127.0.0.1:8000/api/sendcsv", json=pd_json)
response = r.json()


excel_df = pd.DataFrame(response['body'])
excel_df = excel_df.sort_values('gruppe')
excel_df = excel_df.fillna("")

# Create an Excel workbook in memory
wb = openpyxl.Workbook()
ws = wb.active

# The excel by default will include rnr and gruppe columns 
# The columns to be shown to user are specified in the command-line option values that come after (-k/--key)
# The column "hu" will be shown if user adds (-c/--colored) option in the command-line and the rows will be tinted according to the column's value 

if keys != None and 'labelIds' in keys:
    not_droppped_columns = ['rnr', 'gruppe', 'colorCode'] 
    if colored:
        not_droppped_columns = ['rnr', 'gruppe', 'hu', 'colorCode']
    else: 
        not_droppped_columns = ['rnr', 'gruppe', 'hu', 'colorCode']

elif colored: not_droppped_columns = ['rnr', 'gruppe', 'hu']
else: not_droppped_columns = ['rnr', 'gruppe']

if keys != None:
    for k in keys:
        for col in excel_df.columns:
            if col != k and col not in not_droppped_columns and col not in keys:
                if col in excel_df.columns:
                    excel_df = excel_df.drop(col, axis=1)
                    
else:
    for col in excel_df.columns:
            if col not in not_droppped_columns:
                excel_df = excel_df.drop(col, axis=1)

def create_excel(excel_df, keys, colored):         
    # colorcode_value = None            
    for r_idx, row in enumerate(dataframe_to_rows(excel_df, index=False, header=True), 1):
        
        if r_idx==1: # first row includes the columns so skip it 
            
            # #  get index values of the following columns
            if 'hu' in excel_df.columns:
                hu_column_index = row.index("hu") + 1
            if keys != None and 'labelIds' in keys:
                colorcode_index = row.index("colorCode")
                labelids_column_index = row.index("labelIds") + 1 
                colorcode_values = row[colorcode_index].split()
                row.pop(colorcode_index) 
            ws.append(row)      
        if r_idx != 1 and keys != None and 'labelIds' in keys:
            row.pop(colorcode_index) 
            
        for c_idx, value in enumerate(row, 1):
            #  check if this is not the column names
            if r_idx!=1 :
                #  store row
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # if labelids is given and the current row iteration matches the position of label
                if keys != None and "labelIds" in keys and c_idx == labelids_column_index and value != "": 
                    
                    if len(colorcode_values) != 0:
                        for cell in ws[r_idx]:
                            tint_fill = PatternFill(start_color=colorcode_values[0][1:], end_color=colorcode_values[0][1:], fill_type="solid")
                            cell.fill = tint_fill
                            
                if colored and r_idx!=1 and c_idx==hu_column_index:
                    given_date = datetime.strptime(f"{value}", "%Y-%m-%d")
                    today_date = datetime.now()
                    months_diff = (today_date.year - given_date.year) * 12 + (today_date.month - given_date.month)
                    if months_diff < 3:
                        for cell in ws[r_idx]:
                            tint_fill = PatternFill(start_color="007500", end_color="007500", fill_type="solid")
                            cell.fill = tint_fill
                    if months_diff > 3 and months_diff < 12:
                        for cell in ws[r_idx]:
                            tint_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                            cell.fill = tint_fill
                    if months_diff > 12:
                        for cell in ws[r_idx]:
                            tint_fill = PatternFill(start_color="b30000", end_color="b30000", fill_type="solid")
                            cell.fill = tint_fill
    # # Save the workbook
    timestamp = datetime.now().date().isoformat()
    wb.save(f"vehicles_{timestamp}.xlsx")
    
create_excel(excel_df=excel_df, keys=keys, colored=colored)