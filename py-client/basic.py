import requests
import json 
import csv 
import pandas as pd
import argparse
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
# endpoint = "https://httpbin.org/200/"
# endpoint = "http://127.0.0.1:8000/api/" # will echo back anything i send to it
endpoint= "https://restcountries.com/v3/all"

'''
This is the client file that takes the key argu
'''

# get_response_data = requests.get(endpoint, json={"query" : "HELLO WORLD!!"}) # API -> MEtho; we are using a web API and it will use a HTTP request

# data = get_response_data.json()

# with open("data.txt", "r", encoding="utf-8") as f:
#   f.write(f'{data}')
#   f.close()

# with open('vehicles.csv', 'r') as file:
#     csv_data = list(csv.reader(file))
    
# url = "https://api.baubuddy.de/index.php/login"

# def get_vehicle_info():

# print(f'{data}') # prnt raw resp
# print(get_response.json())
# HTTP Request -> HTML
# REST API request -> JSON

#  TODO:  get command line arguemnts -k/--keys and -c/--colored



# TODO:  send csv content as post and send the command line arguments as parameters
# csv_content = ''
# with open("vehicles.csv",'r', encoding= 'utf-8') as f:
#     csv_content=f.read()
#     # print(f.read())

# Take command-line arguements 
parser = argparse.ArgumentParser(description="Take an input parameter -k/--keys that can receive an arbitrary amount of string arguments; Take an input parameter -c/--colored that receives a boolean flag and defaults to True")
parser.add_argument('-k', '--keys', nargs='+')      # option that takes a value
parser.add_argument('-c', '--colored', action='store_true')
args = parser.parse_args()

keys = args.keys
colored = args.colored

# Read csv file
pd_csv = pd.read_csv('vehicles.csv', sep=';', encoding='utf-8', keep_default_na=False, na_values=['NaN'])
#  convert the csv data to json data
pd_json = pd_csv.to_json(orient='records', force_ascii=False)



#TODO: to access the formate of respon is {'body' : '[{'gruppe':, 'jdlskj'}]'} so the content of body must be parsed again to jsonhow to map _headersmapping to string in python
r = requests.post(url="http://127.0.0.1:8000/api/sendcsv", json=pd_json)
response = r.json()

# TODO: Create a function that produces excel file

with open("resulttext.txt", 'w') as f:
    f.write(json.dumps(response))

excel_df = pd.DataFrame(response['body'])
excel_df = excel_df.sort_values('gruppe')
excel_df = excel_df.fillna("")

# Create an Excel workbook in memory
wb = openpyxl.Workbook()
ws = wb.active

# Add data from the DataFrame to the workbook
# for row in dataframe_to_rows(excel_df, index=False, header=True):
#     ws.append(row)


# Iterate through the rows to tint specific rows based on the "Age" column value
# columns = excel_json[0].keys()
# header_row = ws.append(columns)

    
# hu_column_index = columns.index("hu") + 1

# for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=hu_column_index, max_col=hu_column_index):
#     for cell in row:
#         age = cell.value
#         if age is not None and age < 25:
#             cell.fill = tint_fill

# for idx, row in excel_df.iterrows():
#     age = row['hu']
#     if colored and age!="":
#         for cell in worksheet["2:" + str(idx + 2)]:
#             for c in cell:
#                 c.fill = tint_fill

# for row_data in excel_json:
#     # row = [row_data[column] for column in columns]
#     ws.append(row)

# print(excel_df.)       

key_indexes = []
# keys = ['lagerort', 'labelIds']
# colored = True

# The excel by default will include rnr and gruppe columns 
# The columns to be shown to user are specified in the command-line option values that come after (-k/--key)
# The column "hu" will be shown if user adds (-c/--colored) option in the command-line and the rows will be tinted according to the column's value 

if keys != None and 'labelIds' in keys:
    not_droppped_columns = ['rnr', 'gruppe', 'colorCode'] 
    if colored:
        not_droppped_columns = ['rnr', 'gruppe', 'hu', 'colorCode'] 

elif colored: not_droppped_columns = ['rnr', 'gruppe', 'hu']
else: not_droppped_columns = ['rnr', 'gruppe']

if keys != None:
    for k in keys:
        for col in excel_df.columns:
            if col != k and col not in not_droppped_columns and col not in keys:
                if col in excel_df.columns:
                    excel_df = excel_df.drop(col, axis=1)
else:
    for col in excel_df.columns:
            if col not in not_droppped_columns:
                excel_df = excel_df.drop(col, axis=1)

def create_excel(excel_df, keys, colored):         
    # colorcode_value = None            
    for r_idx, row in enumerate(dataframe_to_rows(excel_df, index=False, header=True), 1):
        
        if r_idx==1: # first row includes the columns so skip it 
            
            # #  get index values of the following columns
            # rnr_idx = row.index("rnr") + 1
            # gruppe_idx = row.index("gruppe") + 1
            if 'hu' in excel_df.columns:
                hu_column_index = row.index("hu") + 1
            if keys != None and 'labelIds' in keys:
                colorcode_index = row.index("colorCode")
                labelids_column_index = row.index("labelIds") + 1 
                colorcode_values = row[colorcode_index].split()
                row.pop(colorcode_index) 
            ws.append(row)      
        if r_idx != 1 and keys != None and 'labelIds' in keys:
            row.pop(colorcode_index) 
            
        # if len(excel_df) == r_idx:
        #     break       
            # ws.append(row[index,,,,])
                # for i in keys:
                #     key_idx = row.index(i) + 1
                #     key_indexes.append(key_idx)
            #         ws.append(row[key_idx])
        
        for c_idx, value in enumerate(row, 1):
            #  check if this is not the column names
            if r_idx!=1 :
                #  store row
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # if labelids is given and the current row iteration matches the position of label
                if keys != None and "labelIds" in keys and c_idx == labelids_column_index and value != "": 
                    
                    if len(colorcode_values) != 0:
                        for cell in ws[r_idx]:
                            tint_fill = PatternFill(start_color=colorcode_values[0][1:], end_color=colorcode_values[0][1:], fill_type="solid")
                            cell.fill = tint_fill
                            
                if colored and r_idx!=1 and c_idx==hu_column_index and value != "": # TODO: PLEASE REMOVE "value != """ THIS AFTER THEY REPLY TO YOUR EMAIL ABOUT THE FILTERING
                    # print(value)
                    # hu_val = value[hu_column_index]
                    given_date = datetime.strptime(f"{value}", "%Y-%m-%d")
                    today_date = datetime.now()
                    months_diff = (today_date.year - given_date.year) * 12 + (today_date.month - given_date.month)
                    if months_diff < 3:
                        for cell in ws[r_idx]:
                            tint_fill = PatternFill(start_color="007500", end_color="007500", fill_type="solid")
                            cell.fill = tint_fill
                    if months_diff > 3 and months_diff < 12:
                        for cell in ws[r_idx]:
                            tint_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                            cell.fill = tint_fill
                    if months_diff > 12:
                        for cell in ws[r_idx]:
                            tint_fill = PatternFill(start_color="b30000", end_color="b30000", fill_type="solid")
                            cell.fill = tint_fill
    # # Save the workbook
    timestamp = datetime.now().date().isoformat()
    wb.save(f"vehicles_{timestamp}.xlsx")
    
create_excel(excel_df=excel_df, keys=keys, colored=colored)


'''
The client task: 
Transmits a CSV to a REST-API (s. Server-section below), handles the response and generates an Excel-File taking the input parameters into account.

Take an input parameter -k/--keys that can receive an arbitrary amount of string arguments

Take an input parameter -c/--colored that receives a boolean flag and defaults to True

Transmit CSV containing vehicle information to the POST Call of the server (example data: vehicles.csv)

Convert the servers response into an excel file that contains all resources and make sure that:

Rows are sorted by response field gruppe
Columns always contain rnr field
TODO: Only keys that match the input arguments are considered as additional columns (i.e. when the script is invoked with kurzname and info, print two extra columns)

If labelIds are given and at least one colorCode could be resolved, use the first colorCode to tint the cell's text (if labelIds is given in -k)

If the -c flag is True, color each row depending on the following logic:
If hu is not older than 3 months --> green (#007500)
If hu is not older than 12 months --> orange (#FFA500)
If hu is older than 12 months --> red (#b30000)
The file should be named vehicles_{current_date_iso_formatted}.xlsx
'''